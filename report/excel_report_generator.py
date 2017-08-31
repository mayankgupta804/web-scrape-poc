from openpyxl import Workbook

from config.properties import Properties
from mongo.mongodb import MongoDB
from utility.counter import Counter


class ExcelReport:
    def __init__(self, mongod):
        self.wb = Workbook()
        self.mongod = mongod

    def create_workbook(self):
        self.create_summary_sheet()
        self.create_broken_links_sheet()
        self.create_missing_image_sheet()
        self.create_spelling_mistakes_sheet()
        self.wb.save("report.xlsx")

    def create_summary_sheet(self):
        ws = self.wb.active
        crawled_urls_count = self.mongod.get_crawled_urls_count()
        missing_images_count = self.mongod.get_missing_images_count()
        spellings_count = self.mongod.get_spellings_count()
        broken_links_count = self.mongod.get_broken_links_count()
        data = [["Base Url", Properties.home_page],
                ["Crawled Urls", str(crawled_urls_count)],
                ["Broken Links", str(broken_links_count) + "/" + str(crawled_urls_count)],
                ["Spelling Mistakes", str(spellings_count) + "/" + str(Counter.total_spellings)],
                ["Missing Images", str(missing_images_count) + "/" + str(Counter.total_images)]
                ]
        for row in data:
            ws.append(row)

    def create_broken_links_sheet(self):
        ws = self.wb.create_sheet("BrokenLinks", 1)
        data = [["Url", "Status Code", "Status"]] + self.mongod.get_all_broken_links()
        for row in data:
            ws.append(row)

    def create_missing_image_sheet(self):
        ws = self.wb.create_sheet("MissingImages", 2)
        data = [["Url", "Status Code", "Status"]] + \
               self.mongod.get_all_missing_images()
        for row in data:
            ws.append(row)

    def create_spelling_mistakes_sheet(self):
        ws = self.wb.create_sheet("SpellingMistakes", 3)
        data = [["Word", "Count", "Page"]] + \
               self.mongod.get_all_spellings()
        for row in data:
            ws.append(row)


if __name__ == "__main__":
    mongod = MongoDB().open("Crawler1504075609")
    ExcelReport(mongod).create_workbook()
